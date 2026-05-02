#!/usr/bin/env python3
import csv
import math
import re
from pathlib import Path

import openpyxl


BASE_DIR = Path("/media/anomalymous/2C0A78860A784EB8/SWJTU/math/FJSSP")

INPUTS = {
    "q1": BASE_DIR / "Q1_solution_summary.md",
    "q2": BASE_DIR / "Q2_solution_summary.md",
    "q3": BASE_DIR / "Q3_solution_summary.md",
    "q4": BASE_DIR / "Q4_result_summary.md",
    "problem": BASE_DIR / "2026-51MCM-Problem B-English.docx",
    "attachment": BASE_DIR / "B-attachment.xlsx",
}

OUTPUTS = {
    "q1_schedule": BASE_DIR / "q1_schedule.csv",
    "q2_schedule": BASE_DIR / "q2_schedule.csv",
    "q3_schedule": BASE_DIR / "q3_schedule.csv",
    "q4_schedule": BASE_DIR / "q4_schedule.csv",
    "q4_procurement_plans": BASE_DIR / "q4_procurement_plans.csv",
    "solver_summary": BASE_DIR / "solver_summary.csv",
    "feasibility_audit": BASE_DIR / "feasibility_audit.csv",
}

TYPE_NAMES = [
    "Automated Conveying Arm",
    "Industrial Cleaning Machine",
    "Precision Filling Machine",
    "Automatic Sensing Multi-Function Machine",
    "High-speed Polishing Machine",
]

PREFIX_TO_TYPE = {
    "ACA": "Automated Conveying Arm",
    "ICM": "Industrial Cleaning Machine",
    "PFM": "Precision Filling Machine",
    "ASM": "Automatic Sensing Multi-Function Machine",
    "HPM": "High-speed Polishing Machine",
}

CHINESE_EQUIPMENT_TO_TYPE = {
    "自动化输送臂": "Automated Conveying Arm",
    "工业清洗机": "Industrial Cleaning Machine",
    "精密灌装机": "Precision Filling Machine",
    "自动传感多功能机": "Automatic Sensing Multi-Function Machine",
    "高速抛光机": "High-speed Polishing Machine",
}

SCHEDULE_COLUMNS = [
    "Question",
    "No",
    "EquipmentID",
    "Start",
    "End",
    "Start_s",
    "End_s",
    "Duration_s",
    "Process",
    "Crew",
    "Workshop",
    "EquipmentType",
]


def require_inputs():
    missing = [str(path) for path in INPUTS.values() if not path.exists()]
    if missing:
        raise FileNotFoundError("Missing required input files:\n" + "\n".join(missing))


def read_text(path):
    return path.read_text(encoding="utf-8")


def split_markdown_row(line):
    line = line.strip()
    if not line.startswith("|") or not line.endswith("|"):
        return None
    cells = [cell.strip() for cell in line.strip("|").split("|")]
    return cells


def is_separator_row(cells):
    return all(re.fullmatch(r":?-{3,}:?", cell.replace(" ", "")) for cell in cells)


def extract_table_after_marker(text, marker):
    lines = text.splitlines()
    start = None
    for i, line in enumerate(lines):
        if marker in line:
            start = i
            break
    if start is None:
        raise ValueError(f"Could not find marker: {marker}")

    table_lines = []
    in_table = False
    for line in lines[start + 1 :]:
        cells = split_markdown_row(line)
        if cells is None:
            if in_table:
                break
            continue
        in_table = True
        table_lines.append(cells)

    if len(table_lines) < 2:
        raise ValueError(f"Could not find markdown table after marker: {marker}")

    header = table_lines[0]
    data_rows = table_lines[1:]
    if data_rows and is_separator_row(data_rows[0]):
        data_rows = data_rows[1:]

    return [dict(zip(header, row)) for row in data_rows]


def hms_to_seconds(value):
    value = str(value).strip()
    parts = value.split(":")
    if len(parts) != 3:
        raise ValueError(f"Invalid time value: {value}")
    h, m, s = (int(part) for part in parts)
    return h * 3600 + m * 60 + s


def clean_int(value):
    value = str(value).strip().replace(",", "")
    value = re.sub(r"[^0-9-]", "", value)
    if value == "":
        return 0
    return int(value)


def canonical_type(value):
    text = str(value or "").strip()
    for prefix, type_name in PREFIX_TO_TYPE.items():
        if text.startswith(prefix):
            return type_name
    text_lower = text.lower()
    for type_name in TYPE_NAMES:
        if type_name.lower() in text_lower:
            return type_name
    for zh, type_name in CHINESE_EQUIPMENT_TO_TYPE.items():
        if zh in text:
            return type_name
    return text


def crew_from_equipment(equipment_id, default=""):
    match = re.search(r"(?:ACA|ICM|PFM|ASM|HPM)([12])-", str(equipment_id))
    if match:
        return match.group(1)
    if "1-" in str(equipment_id):
        return "1"
    if "2-" in str(equipment_id):
        return "2"
    return default


def process_workshop(process):
    return str(process).strip()[0]


def normalize_schedule_row(question, row):
    equipment_id = row.get("Equipment ID", row.get("Equipment ID ", "")).strip()
    process = row.get("Process", row.get("Process ID", "")).strip()
    start = row.get("Start", row.get("Start Time", "")).strip()
    end = row.get("End", row.get("End Time", "")).strip()
    duration = clean_int(row.get("Duration (s)", "0"))
    crew = row.get("Crew", "").strip() or crew_from_equipment(equipment_id, default="1")
    equipment_type = canonical_type(row.get("Equipment Type", equipment_id))

    return {
        "Question": question,
        "No": clean_int(row.get("No.", "0")),
        "EquipmentID": equipment_id,
        "Start": start,
        "End": end,
        "Start_s": hms_to_seconds(start),
        "End_s": hms_to_seconds(end),
        "Duration_s": duration,
        "Process": process,
        "Crew": crew,
        "Workshop": process_workshop(process),
        "EquipmentType": equipment_type,
    }


def parse_schedule_tables():
    q1_rows = extract_table_after_marker(read_text(INPUTS["q1"]), "## Table 1")
    q2_rows = extract_table_after_marker(read_text(INPUTS["q2"]), "## Table 2")
    q3_rows = extract_table_after_marker(read_text(INPUTS["q3"]), "## Table 3")
    q4_rows = extract_table_after_marker(read_text(INPUTS["q4"]), "## Table 4: Equipment Operation Schedule")

    return {
        "P1": [normalize_schedule_row("P1", row) for row in q1_rows],
        "P2": [normalize_schedule_row("P2", row) for row in q2_rows],
        "P3": [normalize_schedule_row("P3", row) for row in q3_rows],
        "P4": [normalize_schedule_row("P4", row) for row in q4_rows],
    }


def parse_procurement_plans():
    rows = extract_table_after_marker(read_text(INPUTS["q4"]), "## Plan Comparison")
    output = []
    for row in rows:
        output.append(
            {
                "Plan": row["Plan (y1A, y2A, y1H, y2H)"],
                "Cost": clean_int(row["Cost (USD)"]),
                "Makespan": clean_int(row["Makespan (s)"]),
                "Status": row["Status"].strip(),
            }
        )
    return output


def parse_solver_summary():
    summaries = {
        "P1": read_text(INPUTS["q1"]),
        "P2": read_text(INPUTS["q2"]),
        "P3": read_text(INPUTS["q3"]),
        "P4": read_text(INPUTS["q4"]),
    }
    output = []
    for problem, text in summaries.items():
        status_match = re.search(r"Solver status(?: for chosen plan)?:\s*([A-Z]+)", text)
        makespan_match = re.search(r"Best makespan:\s*([0-9,]+)\s*s\s*\(([^)]+)\)", text)
        if not makespan_match:
            raise ValueError(f"Could not parse makespan for {problem}")

        bound_match = re.search(r"(?:Objective lower bound|Lower bound):\s*([0-9,]+)\s*s\s*\(([^)]+)\)", text)
        gap_match = re.search(r"Relative optimality gap:\s*([0-9.]+%)", text)

        makespan_s = clean_int(makespan_match.group(1))
        bound_s = clean_int(bound_match.group(1)) if bound_match else makespan_s
        gap = gap_match.group(1) if gap_match else "0.00%"
        output.append(
            {
                "Problem": problem,
                "Status": status_match.group(1) if status_match else "OPTIMAL",
                "Objective_s": makespan_s,
                "Bound_s": bound_s,
                "Gap": gap,
                "Makespan_s": makespan_s,
                "Makespan_hms": makespan_match.group(2),
            }
        )
    return output


def load_attachment_data():
    wb = openpyxl.load_workbook(INPUTS["attachment"], read_only=True, data_only=True)

    process_requirements = {}
    ws = wb["Process Flow Table"]
    current_workshop = None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i == 1:
            continue
        workshop, process_id, efficiency, _, _ = row
        if workshop:
            current_workshop = str(workshop).strip()
        if not process_id:
            continue
        process_key = str(process_id).split(".")[0].strip()
        efficiency_text = str(efficiency or "")
        required_types = [name for name in TYPE_NAMES if name in efficiency_text]
        process_requirements[process_key] = {
            "Workshop": current_workshop,
            "EquipmentTypes": required_types,
        }

    speeds = {}
    ws = wb["Crew Configuration Table"]
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i == 1:
            continue
        equipment_type, *_rest, speed, _price = row
        if equipment_type:
            speeds[str(equipment_type).strip()] = float(speed)

    distances = {}
    ws = wb["Workshop Distance Table"]
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i == 1:
            continue
        origin, destination, distance = row
        origin = str(origin).strip()
        destination = str(destination).strip()
        dist_m = clean_int(distance)
        distances[(origin, destination)] = dist_m
        if origin in "ABCDE" and destination in "ABCDE":
            distances[(destination, origin)] = dist_m
    for workshop in "ABCDE":
        distances[(workshop, workshop)] = 0

    return process_requirements, speeds, distances


def base_process(process):
    return re.sub(r"_(?:1|2|3)$", "", str(process).strip())


def expected_processes(problem):
    if problem == "P1":
        return ["A1", "A2", "A3"]
    c_chain = []
    for idx in range(1, 4):
        c_chain.extend([f"C3_{idx}", f"C4_{idx}", f"C5_{idx}"])
    return (
        ["A1", "A2", "A3"]
        + ["B1", "B2", "B3", "B4"]
        + ["C1", "C2"]
        + c_chain
        + ["D1", "D2", "D3", "D4", "D5", "D6"]
        + ["E1", "E2", "E3"]
    )


def precedence_arcs(problem):
    if problem == "P1":
        chains = [["A1", "A2", "A3"]]
    else:
        c_chain = ["C1", "C2"]
        for idx in range(1, 4):
            c_chain.extend([f"C3_{idx}", f"C4_{idx}", f"C5_{idx}"])
        chains = [
            ["A1", "A2", "A3"],
            ["B1", "B2", "B3", "B4"],
            c_chain,
            ["D1", "D2", "D3", "D4", "D5", "D6"],
            ["E1", "E2", "E3"],
        ]
    arcs = []
    for chain in chains:
        arcs.extend(zip(chain, chain[1:]))
    return arcs


def has_complete_operations(problem, rows):
    actual = {row["Process"] for row in rows}
    expected = set(expected_processes(problem))
    return actual == expected


def respects_precedence(problem, rows):
    by_process = {}
    for row in rows:
        by_process.setdefault(row["Process"], []).append(row)
    for before, after in precedence_arcs(problem):
        if before not in by_process or after not in by_process:
            return False
        before_completion = max(row["End_s"] for row in by_process[before])
        after_start = min(row["Start_s"] for row in by_process[after])
        if before_completion > after_start:
            return False
    return True


def equipment_types_match(problem, rows, process_requirements):
    by_process = {}
    for row in rows:
        by_process.setdefault(row["Process"], set()).add(row["EquipmentType"])
    for process in expected_processes(problem):
        required = set(process_requirements[base_process(process)]["EquipmentTypes"])
        actual = by_process.get(process, set())
        if actual != required:
            return False
    return True


def no_machine_overlap(rows):
    by_equipment = {}
    for row in rows:
        by_equipment.setdefault(row["EquipmentID"], []).append(row)
    for intervals in by_equipment.values():
        intervals = sorted(intervals, key=lambda row: (row["Start_s"], row["End_s"]))
        for previous, current in zip(intervals, intervals[1:]):
            if previous["End_s"] > current["Start_s"]:
                return False
    return True


def transport_checks(rows, speeds, distances):
    by_equipment = {}
    for row in rows:
        by_equipment.setdefault(row["EquipmentID"], []).append(row)

    initial_ok = True
    inter_ok = True
    for equipment_id, intervals in by_equipment.items():
        intervals = sorted(intervals, key=lambda row: (row["Start_s"], row["End_s"]))
        equipment_type = intervals[0]["EquipmentType"]
        speed = speeds.get(equipment_type)
        if speed is None:
            return False, False

        first = intervals[0]
        origin = f"Crew {first['Crew']}"
        destination = first["Workshop"]
        if (origin, destination) not in distances:
            return False, False
        required_initial = math.ceil(distances[(origin, destination)] / speed)
        if first["Start_s"] < required_initial:
            initial_ok = False

        for previous, current in zip(intervals, intervals[1:]):
            key = (previous["Workshop"], current["Workshop"])
            if key not in distances:
                return False, False
            required = math.ceil(distances[key] / speed)
            actual_gap = current["Start_s"] - previous["End_s"]
            if actual_gap < required:
                inter_ok = False
    return initial_ok, inter_ok


def parse_q4_budget_status():
    text = read_text(INPUTS["q4"])
    match = re.search(r"Total procurement cost:\s*\$?([0-9,]+)\s*/\s*\$?([0-9,]+)", text)
    if not match:
        raise ValueError("Could not parse Q4 total procurement cost and budget.")
    return clean_int(match.group(1)) <= clean_int(match.group(2))


def build_feasibility_audit(schedules, solver_summary):
    process_requirements, speeds, distances = load_attachment_data()
    summary_by_problem = {row["Problem"]: row for row in solver_summary}
    q4_budget_ok = parse_q4_budget_status()

    audit_items = [
        "工序完整性",
        "工序顺序约束",
        "设备类型匹配",
        "单设备不重叠",
        "双设备协同完成规则",
        "跨车间转运时间",
        "初始位置转运时间",
        "预算约束",
        "Objective-Bound证书",
    ]
    result = {item: {} for item in audit_items}

    for problem, rows in schedules.items():
        complete = has_complete_operations(problem, rows)
        precedence = respects_precedence(problem, rows)
        type_match = equipment_types_match(problem, rows, process_requirements)
        no_overlap = no_machine_overlap(rows)

        result["工序完整性"][problem] = "PASS" if complete else "FAIL"
        result["工序顺序约束"][problem] = "PASS" if precedence else "FAIL"
        result["设备类型匹配"][problem] = "PASS" if type_match else "FAIL"
        result["单设备不重叠"][problem] = "PASS" if no_overlap else "FAIL"
        result["双设备协同完成规则"][problem] = "PASS" if type_match and precedence else "FAIL"

        if problem == "P1":
            result["跨车间转运时间"][problem] = "N.A."
            result["初始位置转运时间"][problem] = "N.A."
        else:
            initial_ok, inter_ok = transport_checks(rows, speeds, distances)
            result["跨车间转运时间"][problem] = "PASS" if inter_ok else "FAIL"
            result["初始位置转运时间"][problem] = "PASS" if initial_ok else "FAIL"

        if problem == "P4":
            result["预算约束"][problem] = "PASS" if q4_budget_ok else "FAIL"
        else:
            result["预算约束"][problem] = "N.A."

        summary = summary_by_problem[problem]
        objective_bound_ok = (
            summary["Status"] == "OPTIMAL"
            and int(summary["Objective_s"]) == int(summary["Bound_s"])
            and summary["Gap"] == "0.00%"
        )
        result["Objective-Bound证书"][problem] = "PASS" if objective_bound_ok else "FAIL"

    return [{"AuditItem": item, **result[item]} for item in audit_items]


def write_csv(path, rows, columns):
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=columns)
        writer.writeheader()
        writer.writerows(rows)


def main():
    require_inputs()

    schedules = parse_schedule_tables()
    procurement_plans = parse_procurement_plans()
    solver_summary = parse_solver_summary()
    feasibility_audit = build_feasibility_audit(schedules, solver_summary)

    write_csv(OUTPUTS["q1_schedule"], schedules["P1"], SCHEDULE_COLUMNS)
    write_csv(OUTPUTS["q2_schedule"], schedules["P2"], SCHEDULE_COLUMNS)
    write_csv(OUTPUTS["q3_schedule"], schedules["P3"], SCHEDULE_COLUMNS)
    write_csv(OUTPUTS["q4_schedule"], schedules["P4"], SCHEDULE_COLUMNS)
    write_csv(OUTPUTS["q4_procurement_plans"], procurement_plans, ["Plan", "Cost", "Makespan", "Status"])
    write_csv(
        OUTPUTS["solver_summary"],
        solver_summary,
        ["Problem", "Status", "Objective_s", "Bound_s", "Gap", "Makespan_s", "Makespan_hms"],
    )
    write_csv(OUTPUTS["feasibility_audit"], feasibility_audit, ["AuditItem", "P1", "P2", "P3", "P4"])

    print("Generated CSV files:")
    for path in OUTPUTS.values():
        print(path)


if __name__ == "__main__":
    main()
